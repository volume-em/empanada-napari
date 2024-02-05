from napari_plugin_engine import napari_hook_implementation
from magicgui import magicgui
import napari
from napari import Viewer
from napari.layers import Labels
import pandas as pd
import os
from openpyxl import Workbook
import numpy as np
import dask.array as da
import itertools


def save_label_lists(label_type, class_names, label_queue, save_dir, labels_layer, plane=None):
    if label_type == 'Current image':
        for class_name in class_names.values():
            if plane == 'null':
                filename = f'{class_name}_label_ids.xlsx'
                sheet_name = labels_layer.name
            else:
                filename = f'{class_name}_image_{plane}_label_ids.xlsx'
                current_image = plane
                sheet_name = f'Image {current_image}'
            file_path = os.path.join(save_dir, filename)
            if os.path.exists(file_path):
                new_filename = f'{class_name}_image_{plane}_label_ids_updated.xlsx'
                file_path = os.path.join(save_dir, new_filename)
            workbook = Workbook()
            sheet = workbook.create_sheet(title=sheet_name)
            sheet['A1'] = 'Label ID'
            for class_id, label_ids in label_queue.items():
                curr_class_name = class_names[class_id]
                if curr_class_name == class_name:
                    for row_num, label_id in enumerate(label_ids, start=2):
                        sheet.cell(row=row_num, column=1, value=label_id)

                workbook.save(file_path)
            try:
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)
                workbook.save(file_path)
            except:
                pass
                # print(f'Saved Excel file for class {class_id} ({class_name}) to {file_path}')

    elif label_type == '3D volume or z-stack':
        sheet_name = labels_layer.name
        workbook = Workbook()
        sheet = workbook.create_sheet(title=sheet_name)
        sheet['A1'] = 'Label ID'
        for class_id, label_ids in label_queue.items():
            class_name = class_names[class_id]
            for row_num, label_id in enumerate(label_ids, start=2):
                sheet.cell(row=row_num, column=1, value=label_id)
            filename = f'{class_name}_volume_label_ids.xlsx'
            file_path = os.path.join(save_dir, filename)
            if os.path.exists(file_path):
                new_filename = f'{class_name}_volume_label_ids_updated.xlsx'
                file_path = os.path.join(save_dir, new_filename)
            workbook.save(file_path)

        try:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
            workbook.save(file_path)
        except:
            pass


def create_xlsx_from_label_queue_list(class_names, label_queues_list, save_dir, labels):
    for class_name in class_names.values():
        filename = f'{class_name}_patch_label_ids.xlsx'
        file_path = os.path.join(save_dir, filename)
        if os.path.exists(file_path):
            new_filename = f'{class_name}_patch_label_ids_updated.xlsx'
            file_path = os.path.join(save_dir, new_filename)
        workbook = Workbook()

        for slice_num in range(labels.shape[0]):
            label_queue = label_queues_list[slice_num]
            for class_id, label_ids in label_queue.items():
                if class_names[class_id] == class_name:
                    sheet_name = f'Image {slice_num}'

                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        # Clear the existing contents of the sheet
                        sheet.delete_rows(2, sheet.max_row)
                    else:
                        # Create a new sheet
                        sheet = workbook.create_sheet(title=sheet_name)
                    sheet['A1'] = 'Label ID'

                    for row_num, label_id in enumerate(label_ids, start=1):
                        sheet.cell(row=row_num + 1, column=1, value=label_id)
        try:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
            workbook.save(file_path)
        except:
            pass
        workbook.save(file_path)


def count_labels(label_values, label_divisor):
    label_queue = {}
    if label_divisor == 0:
        label_queue[1] = label_values.tolist()
        return label_queue, [1]

    class_ids = np.unique(np.floor_divide(label_values, label_divisor)).tolist()
    for ci in class_ids:
        min_id = ci * label_divisor + 1
        max_id = (ci + 1) * label_divisor
        label_ids = label_values[np.logical_and(label_values >= min_id, label_values < max_id)]
        label_queue[ci] = label_ids.tolist()

    return label_queue, class_ids


def label_counter_widget():
    label_params = {
        'Current image': 'Current image',
        '2D patches': '2D patches',
        '3D volume or z-stack': '3D volume or z-stack',
    }

    @magicgui(
        call_button='Count Label IDs (see terminal)',
        layout='vertical',
        label_type=dict(widget_type='RadioButtons', choices=list(label_params.keys()),
                        value=list(label_params.keys())[0], label='Apply to:',
                        tooltip='Calculate number of instance labels'),
        label_text=dict(widget_type='TextEdit', label='Define dataset labels:', value='class_number,class_name',
                        tooltip='Use a separate line for each class. Each line must be {class_number},{class_name (nospaces)}'),
        label_divisor=dict(widget_type='LineEdit', label='Label Divisor', value='0',
                           tooltip='Label divisor that separates objects of different classes.'),

        save_op_head=dict(widget_type='Label', label=f'<h3 text-align="center">Export Label IDs (optional)</h3>',
                          tooltip='Export excel file with label IDs listed by class.'),
        export_xlsx=dict(widget_type='CheckBox', value=False, label='Export list of label IDs (.xlsx file)',
                         tooltip='Export list of label IDs as an excel file.'),
        folder_name=dict(widget_type='LineEdit', value='', label='Folder name'),
        save_dir=dict(widget_type='FileEdit', value='', label='Save directory', mode='d',
                      tooltip='Directory in which to save label counter excel file.'),
    )
    def widget(
            viewer: napari.viewer.Viewer,
            labels_layer: Labels,
            label_type: str,
            label_text: str,
            label_divisor: str,

            save_op_head,
            export_xlsx: bool,
            folder_name: str,
            save_dir: str

    ):
        if export_xlsx:
            # folder_name = f'{labels_layer.name}_label_ids'
            folder_path = os.path.join(save_dir, folder_name)
            # Create the save directory if it doesn't exist
            save_dir = folder_path
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)

        label_divisor = int(label_divisor)
        assert label_divisor > -1, "Label divisor must be a non-negative integer!"

        labels = labels_layer.data

        if labels.ndim > 2 and label_type == 'Current image':
            # assert viewer.dims.order[0] == 0, "Must be viewing axis 0 (xy plane)!"
            plane = viewer.dims.current_step[0]
        else:
            plane = 'null'

        if plane != 'null':
            labels = labels[plane]

        class_names = {}
        for seg_class in label_text.split():
            class_id, class_name = seg_class.split(',')
            class_num = class_id.strip()
            class_name = class_name.strip()
            class_names[int(class_num)] = class_name

        print(f'Class names: {class_names}')

        if isinstance(labels, da.Array):
            label_values = []
            for inds in itertools.product(*map(range, labels.blocks.shape)):
                chunk = labels.blocks[inds].compute()
                label_values.append(np.unique(chunk)[1:])
            label_values = np.concatenate(label_values)
        else:
            label_values = np.unique(labels)[1:]

        if label_type == 'Current image':
            label_queue, class_ids = count_labels(label_values, label_divisor)
            if label_queue and class_ids:
                class_ids = np.array(class_ids)
                has_labels = np.isin(class_ids, list(label_queue.keys()))
                valid_class_ids = class_ids[has_labels]
                valid_label_lists = [np.unique(label_queue[class_id]) for class_id in valid_class_ids]
                valid_label_counts = [len(label_list) for label_list in valid_label_lists]

                for class_id, label_list, label_count in zip(valid_class_ids, valid_label_lists, valid_label_counts):
                    if label_count > 0:
                        print(f'Label IDs in class {class_id} ({class_names[class_id]}): {label_list}')
                        print(f'Total number of label IDs in class {class_id} ({class_names[class_id]}): {label_count}')

                    else:
                        print(f'No label IDs in class {class_id} ({class_names[class_id]}) found in current slice!')

            if export_xlsx:
                os.makedirs(save_dir, exist_ok=True)
                save_label_lists(label_type, class_names, label_queue, save_dir, labels_layer, plane)
                print(f'Saved Excel file to {save_dir}')

        elif label_type == '2D patches':
            class_ids_list = []
            label_queues_list = []

            for slice_num in range(labels.shape[0]):
                label_slice = labels[slice_num]

                if isinstance(label_slice, da.Array):
                    slice_labels = []
                    for sl in itertools.product(*map(range, label_slice.blocks.shape)):
                        chunk = label_slice.blocks[sl].compute()
                        slice_labels.append(np.unique(chunk)[1:])
                    slice_labels = np.concatenate(slice_labels)
                else:
                    slice_labels = np.unique(label_slice)[1:]

                label_queue, class_ids = count_labels(slice_labels, label_divisor)
                label_queues_list.append(label_queue)
                class_ids_list.append(class_ids)

                if label_queue and class_ids:
                    for class_id in class_ids:
                        if class_id in label_queue:
                            label_list = np.unique((label_queue[class_id]))
                            if len(label_list) > 0:
                                print(f'Total number of label IDs in class {class_id} ({class_names[class_id]}) in image {slice_num}:', len(label_list))
                            else:
                                print(f'No label IDs in class {class_id} ({class_names[class_id]}) in image {slice_num}!')

            if export_xlsx:
                os.makedirs(save_dir, exist_ok=True)
                create_xlsx_from_label_queue_list(class_names, label_queues_list, save_dir, labels)
                print(f'Saved Excel file to {save_dir}')

        elif label_type == '3D volume or z-stack':
            label_queue, class_ids = count_labels(label_values, label_divisor)

            if label_queue and class_ids:
                for class_id in class_ids:
                    if class_id in label_queue:
                        label_list = np.unique((label_queue[class_id]))
                        if len(label_list) > 0:
                            print(f'Total number of label IDs in class {class_id} ({class_names[class_id]}) in volume:', len(label_list))

                        else:
                            print(f'No label IDs in class {class_id} ({class_names[class_id]}) in volume!')

            if export_xlsx:
                os.makedirs(save_dir, exist_ok=True)
                save_label_lists(label_type, class_names, label_queue, save_dir, labels_layer)
                print(f'Saved Excel file to {save_dir}')

    return widget


@napari_hook_implementation(specname='napari_experimental_provide_dock_widget')
def label_count_widget():
    return label_counter_widget, {'name': 'Count Labels'}
